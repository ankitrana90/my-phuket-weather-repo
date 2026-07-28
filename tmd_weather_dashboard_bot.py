"""
TMD Phuket Weather -> Raw Excel Log + Advanced Physics Apple Weather Matrix (Single Script).

Excel (Phuket_Weather.xlsx) logs raw TMD observations alongside real-time condition descriptions
fetched from Open-Meteo API. The generated HTML dashboard (Phuket_Weather_Dashboard.html) features:
  - Dates in DESCENDING order (newest on top).
  - Clean cells (zeros hidden).
  - 8-Slot real-time condition emoji row beneath each Date header (blank space if no data).
  - Physics-based fluid dynamics (gravity wave propagation, sloshing, and splash mechanics).
  - Color gradient scaling from light blue to dark navy at >= 8mm.
  - Rain particles tilted dynamically up to 60 degrees based on wind speed.
"""

import argparse
import json
import math
import os
import time
import xml.etree.ElementTree as ET
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import requests
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# ----------------------------------------------------------------------------
# CONFIG
# ----------------------------------------------------------------------------

TMD_UID = os.environ.get("TMD_UID", "api")
TMD_UKEY = os.environ.get("TMD_UKEY", "api12345")
TMD_URL = "https://data.tmd.go.th/api/Weather3Hours/V2/"

# Open-Meteo Free Weather API for Phuket (Lat: 7.8804, Lon: 98.3923)
OPEN_METEO_URL = "https://api.open-meteo.com/v1/forecast?latitude=7.8804&longitude=98.3923&current_weather=true"

TARGET_STATION_NAMES = {"PHUKET AIRPORT", "PHUKET"}
TARGET_WMO_NUMBERS = {"48565", "48564"}

VISIBILITY_TAG_CANDIDATES = ["VisibilityLand", "LandVisibility", "Visibility", "Vis"]

OUTPUT_DIR = Path(".")
EXCEL_PATH = OUTPUT_DIR / "Phuket_Weather.xlsx"
HTML_PATH = OUTPUT_DIR / "Phuket_Weather_Dashboard.html"

DATA_SHEET = "Phuket Weather"

LEGACY_SHEET_NAMES_TO_DROP = {"Dashboard", "Lists"}
LEGACY_SHEET_PREFIXES_TO_DROP = ("_pivot_",)

HEADERS = ["Station", "WmoStationNumber", "DateTime", "WindSpeed", "Rainfall_mm", "Rainfall24Hr_mm", "LandVisibility", "Condition"]

RAINFALL_CAP_MM = 8

STATION_CODES = ["AIRPORT", "PHUKET", "COMBINED"]
STATION_LABELS = {
    "AIRPORT": "PHUKET AIRPORT",
    "PHUKET": "PHUKET",
    "COMBINED": "PHUKET (COMBINED)",
}
STATION_DISPLAY = {
    "AIRPORT": "Phuket Airport",
    "PHUKET": "Phuket",
    "COMBINED": "Phuket (Combined)",
}

METRICS = [
    ("Rain", "Rainfall_mm", "Rainfall (mm)"),
    ("Wind", "WindSpeed", "Wind Speed (km/h)"),
    ("Vis", "LandVisibility", "Land Visibility (km)"),
]

INTERVAL_HOURS = 3

# WMO Weather Interpretation Codes (Open-Meteo)
WMO_CODE_MAP = {
    0: ("Clear sky", "☀️"),
    1: ("Mainly clear", "🌤️"),
    2: ("Partly cloudy", "⛅"),
    3: ("Overcast", "☁️"),
    45: ("Fog", "🌫️"),
    48: ("Depositing rime fog", "🌫️"),
    51: ("Light drizzle", "🌦️"),
    53: ("Moderate drizzle", "🌦️"),
    55: ("Dense drizzle", "🌧️"),
    61: ("Slight rain", "🌦️"),
    63: ("Moderate rain", "🌧️"),
    65: ("Heavy rain", "🌧️"),
    80: ("Slight rain showers", "🌦️"),
    81: ("Moderate rain showers", "🌧️"),
    82: ("Violent rain showers", "⛈️"),
    95: ("Thunderstorm", "⛈️"),
    96: ("Thunderstorm with slight hail", "⛈️"),
    99: ("Thunderstorm with heavy hail", "⛈️"),
}

# ----------------------------------------------------------------------------
# Shared Helpers
# ----------------------------------------------------------------------------

def safe_float(value):
    if value in (None, ""):
        return math.nan
    try:
        return float(value)
    except (TypeError, ValueError):
        return math.nan

def none_if_nan(value):
    return None if (value is None or (isinstance(value, float) and math.isnan(value))) else value

def mean_ignore_nan(values):
    present = [v for v in values if not math.isnan(v)]
    if not present:
        return None
    return round(sum(present) / len(present), 2)

def find_text_any(elem, candidates):
    for tag in candidates:
        found = elem.find(tag)
        if found is not None and found.text is not None and found.text.strip() != "":
            return found.text.strip()
    return None

def fetch_live_weather_condition():
    """Fetches real-time condition description and emoji from Open-Meteo API."""
    try:
        resp = requests.get(OPEN_METEO_URL, timeout=10)
        resp.raise_for_status()
        data = resp.json()
        current = data.get("current_weather", {})
        code = int(current.get("weathercode", 0))
        text, emoji = WMO_CODE_MAP.get(code, ("Variable Conditions", "🌤️"))
        return f"{emoji} {text}"
    except Exception as e:
        print(f"Warning: Could not fetch Open-Meteo condition ({e}). Using fallback.")
        return "🌤️ Variable Conditions"

# ----------------------------------------------------------------------------
# Fetch & Parse
# ----------------------------------------------------------------------------

def fetch_raw_records():
    live_condition = fetch_live_weather_condition()
    resp = requests.get(TMD_URL, params={"uid": TMD_UID, "ukey": TMD_UKEY}, timeout=30)
    resp.raise_for_status()
    root = ET.fromstring(resp.content)

    records = []
    for station in root.iter("Station"):
        name = (station.findtext("StationNameEnglish") or "").strip()
        wmo = (station.findtext("WmoStationNumber") or "").strip()

        if name.upper() not in TARGET_STATION_NAMES and wmo not in TARGET_WMO_NUMBERS:
            continue

        obs = station.find("Observation")
        if obs is None:
            continue

        dt = (obs.findtext("DateTime") or "").strip()
        wind = safe_float(obs.findtext("WindSpeed"))
        rain_elem = obs.find("Rainfall")
        rain = safe_float(rain_elem.text if rain_elem is not None else None)
        rain24_elem = obs.find("Rainfall24Hr")
        rain24 = safe_float(rain24_elem.text if rain24_elem is not None else None)
        visibility = safe_float(find_text_any(obs, VISIBILITY_TAG_CANDIDATES))

        records.append({
            "station": name.upper(),
            "wmo": wmo,
            "datetime": dt,
            "wind_speed": wind,
            "rainfall_mm": rain,
            "rainfall_24hr_mm": rain24,
            "land_visibility": visibility,
            "condition": live_condition
        })
    return records

def build_station_rows(raw_records):
    groups = defaultdict(list)
    for r in raw_records:
        groups[r["datetime"]].append(r)

    rows = []
    for dt in sorted(groups.keys()):
        recs = groups[dt]
        wmo_list = sorted({r["wmo"] for r in recs if r["wmo"]})
        cond = recs[0]["condition"] if recs else "🌤️ Variable Conditions"

        for r in recs:
            rows.append({
                "station": r["station"],
                "wmo": r["wmo"],
                "datetime": dt,
                "wind_speed": none_if_nan(r["wind_speed"]),
                "rainfall_mm": none_if_nan(r["rainfall_mm"]),
                "rainfall_24hr_mm": none_if_nan(r["rainfall_24hr_mm"]),
                "land_visibility": none_if_nan(r["land_visibility"]),
                "condition": r["condition"]
            })

        rows.append({
            "station": STATION_LABELS["COMBINED"],
            "wmo": "+".join(wmo_list),
            "datetime": dt,
            "wind_speed": mean_ignore_nan([r["wind_speed"] for r in recs]),
            "rainfall_mm": mean_ignore_nan([r["rainfall_mm"] for r in recs]),
            "rainfall_24hr_mm": mean_ignore_nan([r["rainfall_24hr_mm"] for r in recs]),
            "land_visibility": mean_ignore_nan([r["land_visibility"] for r in recs]),
            "condition": cond
        })

    return rows

def fetch_records():
    return build_station_rows(fetch_raw_records())

# ----------------------------------------------------------------------------
# Excel Storage Engine
# ----------------------------------------------------------------------------

def load_or_create_workbook():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    if EXCEL_PATH.exists():
        wb = load_workbook(EXCEL_PATH)
        ws = wb[DATA_SHEET] if DATA_SHEET in wb.sheetnames else wb.create_sheet(DATA_SHEET)
        if ws.max_row == 1 and ws.cell(1, 1).value is None:
            ws.append(HEADERS)

        for name in list(wb.sheetnames):
            if name in LEGACY_SHEET_NAMES_TO_DROP or name.startswith(LEGACY_SHEET_PREFIXES_TO_DROP):
                del wb[name]
        wb.active = wb.sheetnames.index(DATA_SHEET)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = DATA_SHEET
        ws.append(HEADERS)

    for col_idx in range(1, len(HEADERS) + 1):
        ws.cell(1, col_idx).font = Font(bold=True)

    return wb, ws

def update_excel(records):
    wb, ws = load_or_create_workbook()

    existing = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        existing.add((row[0], row[2]))

    added = 0
    for r in records:
        key = (r["station"], r["datetime"])
        if key in existing:
            continue
        ws.append([r["station"], r["wmo"], r["datetime"], r["wind_speed"], r["rainfall_mm"], r["rainfall_24hr_mm"], r["land_visibility"], r.get("condition", "")])
        existing.add(key)
        added += 1

    for col_idx in range(1, len(HEADERS) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 20

    wb.save(EXCEL_PATH)
    return added, ws.max_row - 1

# ----------------------------------------------------------------------------
# HTML Dashboard Generator
# ----------------------------------------------------------------------------

def generate_html_dashboard(html_path=HTML_PATH):
    config_json = json.dumps({
        "data_sheet_name": DATA_SHEET,
        "xlsx_filename": EXCEL_PATH.name,
        "metrics": [{"code": c, "column": col, "label": label} for c, col, label in METRICS],
        "stations": [
            {"code": c, "label_raw": STATION_LABELS[c], "label_display": STATION_DISPLAY[c]}
            for c in STATION_CODES
        ],
        "rainfall_cap": RAINFALL_CAP_MM,
    })
    html = _HTML_TEMPLATE.replace("__CONFIG_JSON__", config_json)
    html_path.parent.mkdir(parents=True, exist_ok=True)
    html_path.write_text(html, encoding="utf-8")


_HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Phuket Weather Intelligence</title>
<script src="https://cdn.jsdelivr.net/npm/xlsx@0.18.5/dist/xlsx.full.min.js"></script>
<style>
  :root {
    --bg-light: #F8FAFC;
    --card-bg: rgba(255, 255, 255, 0.85);
    --card-border: 1px solid rgba(226, 232, 240, 0.8);
    --card-shadow: 0 15px 35px -10px rgba(15, 23, 42, 0.05);
    --text-primary: #0F172A;
    --text-secondary: #64748B;
    --accent-blue: #3B82F6;
  }

  * { box-sizing: border-box; margin: 0; padding: 0; }

  body {
    font-family: -apple-system, BlinkMacSystemFont, "SF Pro Display", "SF Pro Text", sans-serif;
    background-color: var(--bg-light);
    color: var(--text-primary);
    min-height: 100vh;
    padding: 32px 24px;
    -webkit-font-smoothing: antialiased;
  }

  .container {
    max-width: 1280px;
    margin: 0 auto;
  }

  header {
    display: flex;
    justify-content: space-between;
    align-items: flex-end;
    margin-bottom: 24px;
  }

  .location-info h1 {
    font-size: 32px;
    font-weight: 600;
    letter-spacing: -0.8px;
    color: var(--text-primary);
  }

  .location-info .meta {
    font-size: 14px;
    color: var(--text-secondary);
    margin-top: 4px;
  }

  .station-select {
    appearance: none;
    background: var(--card-bg);
    border: var(--card-border);
    backdrop-filter: blur(20px);
    color: var(--text-primary);
    padding: 10px 18px;
    border-radius: 18px;
    font-size: 14px;
    font-weight: 500;
    outline: none;
    cursor: pointer;
    box-shadow: var(--card-shadow);
    transition: all 0.2s ease;
  }
  .station-select:hover {
    transform: translateY(-2px);
    border-color: var(--accent-blue);
  }

  /* Hero Display Card */
  .hero-card {
    background: var(--card-bg);
    border: var(--card-border);
    backdrop-filter: blur(25px);
    border-radius: 28px;
    padding: 28px 32px;
    box-shadow: var(--card-shadow);
    display: flex;
    justify-content: space-between;
    align-items: center;
    margin-bottom: 28px;
  }

  .hero-left {
    display: flex;
    align-items: center;
    gap: 20px;
  }

  .hero-emoji {
    font-size: 48px;
    line-height: 1;
  }

  .hero-val {
    font-size: 44px;
    font-weight: 300;
    letter-spacing: -1.5px;
    line-height: 1;
  }

  .hero-status {
    font-size: 14px;
    color: var(--text-secondary);
    margin-top: 6px;
  }

  .grid-title {
    font-size: 20px;
    font-weight: 600;
    letter-spacing: -0.4px;
    margin-bottom: 16px;
  }

  .matrix-wrapper {
    background: var(--card-bg);
    border: var(--card-border);
    backdrop-filter: blur(25px);
    border-radius: 24px;
    box-shadow: var(--card-shadow);
    overflow: auto;
    padding: 12px;
  }

  table {
    width: 100%;
    border-collapse: separate;
    border-spacing: 10px;
    font-size: 13px;
    text-align: center;
  }

  th {
    padding: 12px 8px;
    color: var(--text-secondary);
    font-weight: 600;
    font-size: 13px;
    letter-spacing: -0.2px;
  }

  th.date-col { text-align: left; padding-left: 16px; min-width: 170px; }
  th.summary-col { min-width: 240px; text-align: left; padding-left: 16px; }

  td.date-cell {
    text-align: left;
    padding: 12px 14px;
    background: rgba(255, 255, 255, 0.6);
    border-radius: 16px;
    border: 1px solid rgba(226, 232, 240, 0.6);
  }

  .date-title { font-size: 15px; font-weight: 600; color: var(--text-primary); }

  /* 8-Slot Time Emoji Strip Under Date */
  .emoji-slots-strip {
    display: flex;
    gap: 4px;
    margin-top: 8px;
    align-items: center;
  }

  .slot-emoji {
    font-size: 13px;
    width: 16px;
    height: 18px;
    display: flex;
    align-items: center;
    justify-content: center;
  }

  .slot-emoji.empty-emoji {
    opacity: 0;
  }

  /* Interactive Physics Fluid Cell */
  td.time-cell {
    position: relative;
    width: 82px;
    height: 80px;
    background: rgba(255, 255, 255, 0.6);
    border-radius: 16px;
    border: 1px solid rgba(226, 232, 240, 0.6);
    overflow: hidden;
    vertical-align: middle;
    transition: transform 0.2s ease, border-color 0.2s ease;
  }

  td.time-cell:hover {
    transform: scale(1.05);
    border-color: var(--accent-blue);
  }

  .cell-canvas {
    position: absolute;
    inset: 0;
    width: 100%; height: 100%;
    pointer-events: none;
    z-index: 1;
  }

  .cell-content {
    position: relative;
    z-index: 3;
    display: flex;
    flex-direction: column;
    align-items: center;
    justify-content: center;
    height: 100%;
    pointer-events: none;
  }

  .val-num {
    font-size: 15px;
    font-weight: 700;
    letter-spacing: -0.3px;
    text-shadow: 0 1px 2px rgba(255, 255, 255, 0.8);
  }

  .val-unit { font-size: 10px; color: var(--text-secondary); font-weight: 500; }

  td.summary-cell {
    text-align: left;
    padding: 12px 18px;
    background: rgba(59, 130, 246, 0.06);
    border-radius: 16px;
    border: 1px solid rgba(59, 130, 246, 0.15);
  }

  .summary-val { font-size: 17px; font-weight: 600; color: var(--accent-blue); }
  .summary-commentary { font-size: 12px; color: var(--text-secondary); line-height: 1.35; margin-top: 4px; }

  @media (max-width: 768px) {
    .hero-card { flex-direction: column; align-items: flex-start; gap: 16px; }
  }
</style>
</head>
<body>

  <div class="container">
    <header>
      <div class="location-info">
        <h1>Phuket, Thailand</h1>
        <div class="meta" id="headerMeta">Syncing live telemetry...</div>
      </div>
      <select id="stationSelect" class="station-select"></select>
    </header>

    <!-- Hero Display Card -->
    <div class="hero-card">
      <div class="hero-left">
        <div class="hero-emoji" id="heroEmoji">🌦️</div>
        <div>
          <div class="hero-val" id="heroVal">-- mm</div>
          <div class="hero-status" id="heroStatus">Scanning precipitation pattern...</div>
        </div>
      </div>
    </div>

    <div class="grid-title">Time-of-Day Rainfall Matrix (Descending Date Order)</div>
    
    <div class="matrix-wrapper">
      <table id="matrixTable"></table>
    </div>
  </div>

<script>
const CONFIG = __CONFIG_JSON__;
const stationSelect = document.getElementById('stationSelect');
let DATA = null;
let activeCanvasRenderers = [];

CONFIG.stations.forEach(s => {
  const opt = document.createElement('option');
  opt.value = s.code; opt.textContent = s.label_display;
  if (s.code === 'COMBINED') opt.selected = true;
  stationSelect.appendChild(opt);
});

async function autoFetchServerData() {
  try {
    const res = await fetch(CONFIG.xlsx_filename, { cache: 'no-cache' });
    if (!res.ok) throw new Error(`HTTP ${res.status}`);
    const buffer = await res.arrayBuffer();
    
    const wb = XLSX.read(buffer, { type: 'array', cellDates: true });
    const sheet = wb.Sheets[CONFIG.data_sheet_name];
    if (!sheet) throw new Error("Data sheet missing");

    const rows = XLSX.utils.sheet_to_json(sheet, { defval: null });
    DATA = buildDataFromRows(rows);
    
    document.getElementById('headerMeta').textContent = 
      `Most Recent Record: ${DATA.latestDateTime || 'Unknown'} • Rendered ${new Date().toLocaleTimeString([], {hour: '2-digit', minute:'2-digit'})}`;
    
    renderMatrix();
  } catch (err) {
    document.getElementById('headerMeta').textContent = 'Displaying cached telemetry';
  }
}

function pad(n) { return n.toString().padStart(2, '0'); }
function parseDateTime(v) {
  if (!v) return null;
  const d = (v instanceof Date) ? v : new Date(v);
  return isNaN(d.getTime()) ? null : d;
}

function buildDataFromRows(rows) {
  const buckets = {};
  const dateSet = new Set();
  const timeSet = new Set();
  const columns = CONFIG.metrics.map(m => m.column);
  let latestParsedDT = null;
  let conditionMap = {}; // station -> date -> time -> emoji

  rows.forEach(row => {
    const dt = parseDateTime(row.DateTime);
    const station = row.Station;
    if (!dt || !station) return;

    if (!latestParsedDT || dt > latestParsedDT) {
      latestParsedDT = dt;
    }

    const dateStr = `${dt.getFullYear()}-${pad(dt.getMonth() + 1)}-${pad(dt.getDate())}`;
    const timeStr = `${pad(dt.getHours())}:${pad(dt.getMinutes())}`;
    dateSet.add(dateStr); timeSet.add(timeStr);

    // Parse Condition Emoji for exact time slot
    if (row.Condition) {
      const parts = row.Condition.trim().split(' ');
      const emoji = parts[0] || '';
      conditionMap[station] = conditionMap[station] || {};
      conditionMap[station][dateStr] = conditionMap[station][dateStr] || {};
      conditionMap[station][dateStr][timeStr] = emoji;
    }

    buckets[station] = buckets[station] || {};
    buckets[station][dateStr] = buckets[station][dateStr] || {};
    const slot = buckets[station][dateStr][timeStr] =
      buckets[station][dateStr][timeStr] || Object.fromEntries(columns.map(c => [c, []]));

    columns.forEach(col => {
      const v = row[col];
      if (v !== null && v !== undefined && v !== '' && !isNaN(Number(v))) slot[col].push(Number(v));
    });
  });

  const dates = Array.from(dateSet).sort((a, b) => b.localeCompare(a));
  const times = Array.from(timeSet).sort((a, b) => a.localeCompare(b));

  const round2 = x => Math.round(x * 100) / 100;
  const mean = arr => arr.length ? round2(arr.reduce((a, b) => a + b, 0) / arr.length) : null;

  const payload = {};
  CONFIG.metrics.forEach(m => {
    payload[m.code] = {};
    CONFIG.stations.forEach(s => {
      const stationLabel = s.label_raw;
      const grid = dates.map(d => times.map(t => {
        const slot = buckets[stationLabel] && buckets[stationLabel][d] && buckets[stationLabel][d][t];
        return slot ? mean(slot[m.column]) : null;
      }));
      const summary = grid.map(rowVals => {
        const present = rowVals.filter(v => v !== null);
        if (!present.length) return null;
        return m.code === 'Rain' ? round2(present.reduce((a, b) => a + b, 0)) : mean(present);
      });
      payload[m.code][s.code] = { grid, summary };
    });
  });

  const latestFormatted = latestParsedDT 
    ? `${latestParsedDT.getFullYear()}-${pad(latestParsedDT.getMonth() + 1)}-${pad(latestParsedDT.getDate())} ${pad(latestParsedDT.getHours())}:${pad(latestParsedDT.getMinutes())}` 
    : 'N/A';

  return { dates, times, payload, latestDateTime: latestFormatted, conditionMap };
}

/* Fluid Color Palette (Light Blue -> Deep Navy Blue for >= 8mm) */
function getWaterColors(rainVal) {
  const cap = 8.0;
  const ratio = Math.min(rainVal, cap) / cap;

  const r = Math.round(147 + (2 - 147) * ratio);
  const g = Math.round(197 + (21 - 197) * ratio);
  const b = Math.round(253 + (38 - 253) * ratio);
  
  return {
    top: `rgba(${r}, ${g}, ${b}, ${0.55 + 0.35 * ratio})`,
    bottom: `rgba(${Math.max(0, r - 20)}, ${Math.max(0, g - 20)}, ${Math.max(0, b - 20)}, ${0.75 + 0.2 * ratio})`,
    isDark: ratio >= 0.75
  };
}

function evaluateWeather(rain) {
  if (rain >= 25) {
    return { emoji: '⛈️', commentary: 'Torrential downpours with thunder activity.' };
  } else if (rain >= 12) {
    return { emoji: '🌧️', commentary: 'Heavy monsoon rainfall across multiple hours.' };
  } else if (rain >= 4) {
    return { emoji: '🌦️', commentary: 'Passing scattered showers throughout the day.' };
  } else if (rain > 0) {
    return { emoji: '🌤️', commentary: 'Light intermittent drizzle with dry intervals.' };
  } else {
    return { emoji: '☀️', commentary: 'Optimal clear conditions with no rain recorded.' };
  }
}

/* Physics-Based Fluid Sloshing & Splash Canvas Engine */
function attachFluidPhysicsCanvas(canvasId, rainVal, windSpeed) {
  const canvas = document.getElementById(canvasId);
  if (!canvas) return;
  const ctx = canvas.getContext('2d');
  canvas.width = canvas.offsetWidth;
  canvas.height = canvas.offsetHeight;

  if (rainVal <= 0) return;

  const colors = getWaterColors(rainVal);
  const fillRatio = Math.min(rainVal, CONFIG.rainfall_cap) / CONFIG.rainfall_cap;
  const targetWaterHeight = canvas.height * fillRatio;

  const safeWind = Math.min(windSpeed || 0, 50);
  const tiltAngleRad = (safeWind / 50) * (60 * Math.PI / 180);
  const xOffset = Math.sin(tiltAngleRad) * 10;
  const yOffset = Math.cos(tiltAngleRad) * 10;

  const particleCount = Math.min(Math.floor(rainVal * 6) + 3, 30);
  const rainDrops = [];
  for (let i = 0; i < particleCount; i++) {
    rainDrops.push({
      x: Math.random() * canvas.width,
      y: Math.random() * canvas.height,
      speed: Math.random() * 2.5 + 2
    });
  }

  let step = 0;
  const waveAmplitude = Math.min(2 + rainVal * 0.8, 8);
  const sloshSpeed = 0.05 + Math.min(rainVal * 0.01, 0.05);

  const splashes = [];
  const splashCount = rainVal >= 4 ? Math.floor(rainVal) : 0;
  for (let i = 0; i < splashCount; i++) {
    splashes.push({
      x: Math.random() * canvas.width,
      y: canvas.height - targetWaterHeight,
      vx: (Math.random() - 0.5) * 1.5,
      vy: -Math.random() * 2 - 1,
      size: Math.random() * 1.5 + 0.8
    });
  }

  function render() {
    ctx.clearRect(0, 0, canvas.width, canvas.height);
    step += sloshSpeed;

    ctx.strokeStyle = colors.isDark ? 'rgba(255, 255, 255, 0.6)' : 'rgba(59, 130, 246, 0.45)';
    ctx.lineWidth = 1.1;
    ctx.beginPath();
    rainDrops.forEach(p => {
      ctx.moveTo(p.x, p.y);
      ctx.lineTo(p.x - xOffset, p.y + yOffset);
      p.y += p.speed;
      p.x -= (xOffset * 0.12);
      if (p.y > canvas.height) {
        p.y = -8;
        p.x = Math.random() * canvas.width;
      }
    });
    ctx.stroke();

    const baseLine = canvas.height - targetWaterHeight;
    const grad = ctx.createLinearGradient(0, baseLine, 0, canvas.height);
    grad.addColorStop(0, colors.top);
    grad.addColorStop(1, colors.bottom);

    ctx.fillStyle = grad;
    ctx.beginPath();
    ctx.moveTo(0, canvas.height);
    ctx.lineTo(0, baseLine);

    for (let x = 0; x <= canvas.width; x += 4) {
      const y1 = Math.sin(x * 0.08 + step) * waveAmplitude;
      const y2 = Math.cos(x * 0.12 - step * 0.8) * (waveAmplitude * 0.5);
      const y = baseLine + y1 + y2;
      ctx.lineTo(x, y);
    }

    ctx.lineTo(canvas.width, canvas.height);
    ctx.closePath();
    ctx.fill();

    if (splashes.length > 0) {
      ctx.fillStyle = colors.isDark ? 'rgba(255, 255, 255, 0.8)' : 'rgba(147, 197, 253, 0.8)';
      splashes.forEach(sp => {
        ctx.beginPath();
        ctx.arc(sp.x, sp.y, sp.size, 0, Math.PI * 2);
        ctx.fill();

        sp.x += sp.vx;
        sp.y += sp.vy;
        sp.vy += 0.12;

        if (sp.y > canvas.height - targetWaterHeight + 4) {
          sp.x = Math.random() * canvas.width;
          sp.y = canvas.height - targetWaterHeight;
          sp.vy = -Math.random() * 2 - 1;
        }
      });
    }

    const animId = requestAnimationFrame(render);
    activeCanvasRenderers.push(animId);
  }

  render();
}

function renderMatrix() {
  if (!DATA) return;

  activeCanvasRenderers.forEach(id => cancelAnimationFrame(id));
  activeCanvasRenderers = [];

  const station = stationSelect.value;
  const rainData = DATA.payload.Rain[station];
  const windData = DATA.payload.Wind[station];

  const latestRain = rainData.summary[0] !== null ? rainData.summary[0] : 0;
  const latestEval = evaluateWeather(latestRain);

  document.getElementById('heroVal').textContent = `${latestRain} mm`;
  document.getElementById('heroEmoji').textContent = latestEval.emoji;
  document.getElementById('heroStatus').textContent = `Latest Record (${DATA.dates[0]}) • ${latestEval.commentary}`;

  const table = document.getElementById('matrixTable');
  table.innerHTML = '';

  // Header Row
  const thead = document.createElement('tr');
  thead.appendChild(Object.assign(document.createElement('th'), {textContent: 'Date', className: 'date-col'}));
  DATA.times.forEach(t => thead.appendChild(Object.assign(document.createElement('th'), {textContent: t})));
  thead.appendChild(Object.assign(document.createElement('th'), {textContent: 'Daily Remarks', className: 'summary-col'}));
  table.appendChild(thead);

  // Date Rows (Descending Order)
  DATA.dates.forEach((d, i) => {
    const tr = document.createElement('tr');
    
    const totalRain = rainData.summary[i] !== null ? rainData.summary[i] : 0;
    const evalData = evaluateWeather(totalRain);

    // Generate 8-Slot Emojis Strip for Date Cell (Blank space if no data logged)
    const slotEmojis = DATA.times.map(t => {
      const emoji = (DATA.conditionMap[station] && DATA.conditionMap[station][d] && DATA.conditionMap[station][d][t]) || '';
      return emoji ? `<span class="slot-emoji" title="${t}">${emoji}</span>` : `<span class="slot-emoji empty-emoji">&nbsp;</span>`;
    }).join('');

    // Date Cell with 8-slot strip
    const dateTd = document.createElement('td');
    dateTd.className = 'date-cell';
    dateTd.innerHTML = `
      <div class="date-title">${d}</div>
      <div class="emoji-slots-strip">${slotEmojis}</div>
    `;
    tr.appendChild(dateTd);

    // Hourly Cells Beside Date
    DATA.times.forEach((t, j) => {
      const v = rainData.grid[i][j];
      const wSpeed = windData.grid[i][j] || 0;
      const td = document.createElement('td');
      td.className = 'time-cell';

      if (v !== null && v > 0) {
        const canvas = document.createElement('canvas');
        canvas.className = 'cell-canvas';
        canvas.id = `canvas-${i}-${j}`;
        td.appendChild(canvas);

        const content = document.createElement('div');
        content.className = 'cell-content';
        const isDark = v >= 6;
        const numColor = isDark ? '#FFFFFF' : '#0F172A';
        content.innerHTML = `<span class="val-num" style="color:${numColor}">${v}</span><span class="val-unit" style="color:${numColor}">mm</span>`;
        td.appendChild(content);

        setTimeout(() => attachFluidPhysicsCanvas(`canvas-${i}-${j}`, v, wSpeed), 50);
      } else {
        td.innerHTML = `<div class="cell-content"></div>`;
      }

      tr.appendChild(td);
    });

    // Daily Summary Cell
    const sumTd = document.createElement('td');
    sumTd.className = 'summary-cell';
    sumTd.innerHTML = `
      <div class="summary-val">${totalRain > 0 ? totalRain + ' mm Total' : '0 mm'}</div>
      <div class="summary-commentary">${evalData.commentary}</div>
    `;
    tr.appendChild(sumTd);

    table.appendChild(tr);
  });
}

stationSelect.addEventListener('change', renderMatrix);
window.addEventListener('resize', renderMatrix);
window.addEventListener('DOMContentLoaded', autoFetchServerData);
</script>
</body>
</html>
"""

# ----------------------------------------------------------------------------
# Orchestration
# ----------------------------------------------------------------------------

def run_cycle():
    print(f"[{datetime.now().isoformat()}] Fetching TMD feed & Open-Meteo conditions...")
    records = fetch_records()
    print(f"Built {len(records)} row(s) (raw + combined) this cycle.")

    added, total = update_excel(records)
    print(f"Excel Data sheet updated (+Condition column): +{added} new row(s), {total} total.")

    generate_html_dashboard(HTML_PATH)
    print(f"Physics Apple Weather UI generated at {HTML_PATH}")

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--once", action="store_true", help="Run a single cycle and exit")
    args = parser.parse_args()

    if args.once:
        run_cycle()
        return

    while True:
        try:
            run_cycle()
        except Exception as e:
            print(f"Error this cycle: {e}")
        print(f"Sleeping {INTERVAL_HOURS} hour(s)...")
        time.sleep(INTERVAL_HOURS * 3600)

if __name__ == "__main__":
    main()
