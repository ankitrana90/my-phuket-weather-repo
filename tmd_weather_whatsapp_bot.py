"""
TMD Phuket Weather -> raw Excel log + dynamic HTML dashboard (single script).

Excel (Phuket_Weather.xlsx) now stores RAW DATA ONLY -- one "Phuket Weather"
Data sheet, nothing derived. All the dynamic/interactive view lives in a
standalone HTML file (Phuket_Weather_Dashboard.html) instead: pick a Metric
(Rainfall / Wind Speed / Land Visibility) and a Station (Phuket Airport /
Phuket / Phuket Combined) from two dropdowns, and a date x time-of-day grid
updates instantly, client-side, from data embedded in the page (no server
needed). Rendering differs by metric:
  - Rainfall (mm): each cell is a bottom-anchored bar whose HEIGHT represents
    the value, coloured light blue -> dark blue as it approaches an upper cap
    of RAINFALL_CAP_MM (8mm by default). Anything above the cap is drawn as a
    fixed "very dark" colour to flag the overflow, not a further colour scale.
  - Wind Speed (km/h) and Land Visibility (km): plain numbers only -- no
    colour, no bar.
A per-day "Daily Summary" column sums rainfall or averages wind/visibility.

IMPORTANT -- LandVisibility field name is UNVERIFIED:
  TMD's Weather3Hours XML schema was not accessible from this environment to
  confirm the exact tag name, so fetch_raw_records() tries several likely
  candidates (VISIBILITY_TAG_CANDIDATES below) in order and uses whichever
  one is present. After your first successful run, open the Data sheet and
  check whether LandVisibility actually has real numbers -- if it's always
  blank, print one station's raw XML (see dump_one_station_xml() at the
  bottom) and add the real tag name to VISIBILITY_TAG_CANDIDATES.

What it does, every cycle:
  1. Fetches https://data.tmd.go.th/api/Weather3Hours/V2/
  2. Filters stations where StationNameEnglish in ("PHUKET AIRPORT", "PHUKET")
     or WmoStationNumber in ("48565", "48564")
  3. For every timestamp, logs THREE distinct rows to the Data sheet:
       - "PHUKET AIRPORT"     (raw reading from that station)
       - "PHUKET"             (raw reading from that station)
       - "PHUKET (COMBINED)"  (average of whichever of the two reported)
     Re-running the script never duplicates rows (dedup key = Station + DateTime).
  4. Regenerates Phuket_Weather_Dashboard.html from the full Data sheet.

Install:
    pip install requests openpyxl pandas

Usage:
    python tmd_weather_dashboard_bot.py --once
    python tmd_weather_dashboard_bot.py                # loops forever, every INTERVAL_HOURS
"""

import argparse
import json
import math
import os
import time
import xml.etree.ElementTree as ET
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import requests
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


# ----------------------------------------------------------------------------
# CONFIG -- edit these
# ----------------------------------------------------------------------------

# Credentials come from environment variables (set as GitHub Actions secrets
# when hosted -- see .github/workflows/update.yml) so they're never committed
# to the repo. The literal defaults below only apply when you run locally
# without setting the env vars.
TMD_UID = os.environ.get("TMD_UID", "api")            # replace with your real registered TMD uid
TMD_UKEY = os.environ.get("TMD_UKEY", "api12345")      # replace with your real registered TMD ukey
TMD_URL = "https://data.tmd.go.th/api/Weather3Hours/V2/"

TARGET_STATION_NAMES = {"PHUKET AIRPORT", "PHUKET"}
TARGET_WMO_NUMBERS = {"48565", "48564"}

# Tag-name candidates tried in order for land visibility -- see the
# IMPORTANT note above. Add the real tag name here once you've confirmed it.
VISIBILITY_TAG_CANDIDATES = ["VisibilityLand", "LandVisibility", "Visibility", "Vis"]

# Relative to the repo root, so this works identically whether you run it
# locally (from the repo folder) or in GitHub Actions (which checks the repo
# out to its working directory). GitHub Pages then serves these same files.
OUTPUT_DIR = Path(".")
EXCEL_PATH = OUTPUT_DIR / "Phuket_Weather.xlsx"
HTML_PATH = OUTPUT_DIR / "Phuket_Weather_Dashboard.html"

DATA_SHEET = "Phuket Weather"

# Sheet names/prefixes used by an OLDER version of this script (Excel pivot
# helper sheets + a Dashboard sheet). Excel now stores raw data only -- if any
# of these leftover sheets exist in an old workbook, load_or_create_workbook()
# removes them so the file stays clean.
LEGACY_SHEET_NAMES_TO_DROP = {"Dashboard", "Lists"}
LEGACY_SHEET_PREFIXES_TO_DROP = ("_pivot_",)

HEADERS = ["Station", "WmoStationNumber", "DateTime", "WindSpeed", "Rainfall_mm", "Rainfall24Hr_mm", "LandVisibility"]

# Rainfall upper cap (mm) for the HTML heatmap's colour/bar scale. Values at
# or below this fill light-blue -> dark-blue proportionally; values above it
# are shown as a fixed "very dark" colour to flag the cap was exceeded.
RAINFALL_CAP_MM = 8

# Station variants that end up as DISTINCT rows in the Data sheet, and as
# DISTINCT selectable options in the HTML dashboard's Station dropdown.
STATION_CODES = ["AIRPORT", "PHUKET", "COMBINED"]
STATION_LABELS = {
    "AIRPORT": "PHUKET AIRPORT",
    "PHUKET": "PHUKET",
    "COMBINED": "PHUKET (COMBINED)",
}
STATION_DISPLAY = {
    "AIRPORT": "Phuket Airport",
    "PHUKET": "Phuket",
    "COMBINED": "Phuket (Combined)",
}

# Metrics selectable in the Dashboard's Metric dropdown.
# (code, Data-sheet column name, display label)
METRICS = [
    ("Rain", "Rainfall_mm", "Rainfall (mm)"),
    ("Wind", "WindSpeed", "Wind Speed (km/h)"),
    ("Vis", "LandVisibility", "Land Visibility (km)"),
]
METRIC_CODE_TO_COLUMN = {code: col for code, col, _ in METRICS}
METRIC_CODE_TO_LABEL = {code: label for code, _, label in METRICS}

INTERVAL_HOURS = 3


# ----------------------------------------------------------------------------
# Small shared helpers
# ----------------------------------------------------------------------------

def safe_float(value):
    """Converts a value to float; missing/blank/non-numeric becomes NaN
    (not None) so it can be averaged without special-casing."""
    if value in (None, ""):
        return math.nan
    try:
        return float(value)
    except (TypeError, ValueError):
        return math.nan


def none_if_nan(value):
    """Converts NaN back to None so blank Excel cells stay genuinely blank
    (instead of showing the string 'nan')."""
    return None if (value is None or (isinstance(value, float) and math.isnan(value))) else value


def mean_ignore_nan(values):
    present = [v for v in values if not math.isnan(v)]
    if not present:
        return None
    return round(sum(present) / len(present), 2)


def find_text_any(elem, candidates):
    """Tries each candidate child-tag name in order, returns the first
    non-empty text found (or None if none of the candidates are present)."""
    for tag in candidates:
        found = elem.find(tag)
        if found is not None and found.text is not None and found.text.strip() != "":
            return found.text.strip()
    return None


# ----------------------------------------------------------------------------
# 1. Fetch + parse
# ----------------------------------------------------------------------------

def fetch_raw_records():
    """Pulls the raw per-station observations (PHUKET AIRPORT + PHUKET) from
    the TMD feed -- one dict per station, not yet combined."""
    resp = requests.get(TMD_URL, params={"uid": TMD_UID, "ukey": TMD_UKEY}, timeout=30)
    resp.raise_for_status()
    root = ET.fromstring(resp.content)

    records = []
    for station in root.iter("Station"):
        name = (station.findtext("StationNameEnglish") or "").strip()
        wmo = (station.findtext("WmoStationNumber") or "").strip()

        if name.upper() not in TARGET_STATION_NAMES and wmo not in TARGET_WMO_NUMBERS:
            continue

        obs = station.find("Observation")
        if obs is None:
            continue

        dt = (obs.findtext("DateTime") or "").strip()
        wind = safe_float(obs.findtext("WindSpeed"))
        rain_elem = obs.find("Rainfall")
        rain = safe_float(rain_elem.text if rain_elem is not None else None)
        rain24_elem = obs.find("Rainfall24Hr")
        rain24 = safe_float(rain24_elem.text if rain24_elem is not None else None)
        visibility = safe_float(find_text_any(obs, VISIBILITY_TAG_CANDIDATES))

        records.append(
            {
                "station": name.upper(),
                "wmo": wmo,
                "datetime": dt,
                "wind_speed": wind,
                "rainfall_mm": rain,
                "rainfall_24hr_mm": rain24,
                "land_visibility": visibility,
            }
        )

    return records


def build_station_rows(raw_records):
    """For every DateTime, emits THREE distinct rows:
      - the raw PHUKET AIRPORT reading (if reported that timestamp)
      - the raw PHUKET reading (if reported that timestamp)
      - a PHUKET (COMBINED) row averaging whichever of the two reported,
        per numeric field, ignoring missing values.
    """
    groups = defaultdict(list)
    for r in raw_records:
        groups[r["datetime"]].append(r)

    rows = []
    for dt in sorted(groups.keys()):
        recs = groups[dt]
        wmo_list = sorted({r["wmo"] for r in recs if r["wmo"]})

        # Raw, distinct per-station rows.
        for r in recs:
            rows.append(
                {
                    "station": r["station"],  # "PHUKET AIRPORT" or "PHUKET"
                    "wmo": r["wmo"],
                    "datetime": dt,
                    "wind_speed": none_if_nan(r["wind_speed"]),
                    "rainfall_mm": none_if_nan(r["rainfall_mm"]),
                    "rainfall_24hr_mm": none_if_nan(r["rainfall_24hr_mm"]),
                    "land_visibility": none_if_nan(r["land_visibility"]),
                }
            )

        # Combined row (average across whichever stations reported).
        rows.append(
            {
                "station": STATION_LABELS["COMBINED"],
                "wmo": "+".join(wmo_list),
                "datetime": dt,
                "wind_speed": mean_ignore_nan([r["wind_speed"] for r in recs]),
                "rainfall_mm": mean_ignore_nan([r["rainfall_mm"] for r in recs]),
                "rainfall_24hr_mm": mean_ignore_nan([r["rainfall_24hr_mm"] for r in recs]),
                "land_visibility": mean_ignore_nan([r["land_visibility"] for r in recs]),
            }
        )

    return rows


def fetch_records():
    """Public entry point: fetch + build all distinct station rows."""
    return build_station_rows(fetch_raw_records())


def dump_one_station_xml():
    """Debug helper: prints the raw XML for the first matching station, so
    you can confirm the real tag name for land visibility (see the IMPORTANT
    note at the top of this file). Not called automatically -- run manually:
        python -c "from tmd_weather_dashboard_bot import dump_one_station_xml; dump_one_station_xml()"
    """
    resp = requests.get(TMD_URL, params={"uid": TMD_UID, "ukey": TMD_UKEY}, timeout=30)
    resp.raise_for_status()
    root = ET.fromstring(resp.content)
    for station in root.iter("Station"):
        name = (station.findtext("StationNameEnglish") or "").strip()
        if name.upper() in TARGET_STATION_NAMES:
            print(ET.tostring(station, encoding="unicode"))
            return
    print("No matching station found in this fetch.")


# ----------------------------------------------------------------------------
# 2. Excel Data sheet update (append-only, deduped on Station + DateTime)
# ----------------------------------------------------------------------------

def load_or_create_workbook():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    if EXCEL_PATH.exists():
        wb = load_workbook(EXCEL_PATH)
        ws = wb[DATA_SHEET] if DATA_SHEET in wb.sheetnames else wb.create_sheet(DATA_SHEET)
        if ws.max_row == 1 and ws.cell(1, 1).value is None:
            ws.append(HEADERS)

        # Remove any leftover Dashboard/Lists/pivot sheets from an older
        # version of this script -- Excel keeps raw data only now.
        for name in list(wb.sheetnames):
            if name in LEGACY_SHEET_NAMES_TO_DROP or name.startswith(LEGACY_SHEET_PREFIXES_TO_DROP):
                del wb[name]
        wb.active = wb.sheetnames.index(DATA_SHEET)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = DATA_SHEET
        ws.append(HEADERS)

    for col_idx in range(1, len(HEADERS) + 1):
        ws.cell(1, col_idx).font = Font(bold=True)

    return wb, ws


def update_excel(records):
    """Appends new rows, skipping any (Station, DateTime) pair already
    logged so re-running the script never creates duplicates."""
    wb, ws = load_or_create_workbook()

    existing = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        existing.add((row[0], row[2]))  # (Station, DateTime)

    added = 0
    for r in records:
        key = (r["station"], r["datetime"])
        if key in existing:
            continue
        ws.append([r["station"], r["wmo"], r["datetime"], r["wind_speed"], r["rainfall_mm"], r["rainfall_24hr_mm"], r["land_visibility"]])
        existing.add(key)
        added += 1

    for col_idx in range(1, len(HEADERS) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 20

    wb.save(EXCEL_PATH)
    return added, ws.max_row - 1


# ----------------------------------------------------------------------------
# 3. HTML dashboard (standalone file that reads Phuket_Weather.xlsx LIVE,
#    client-side, every time it's opened -- never bakes a data snapshot in)
# ----------------------------------------------------------------------------

def generate_html_dashboard(html_path=HTML_PATH):
    """Writes the self-contained HTML dashboard. It does NOT embed a snapshot
    of the Excel data -- only small, unchanging CONFIG (metric/station
    labels, the rainfall cap, the xlsx filename) -- and uses SheetJS, in the
    browser, to read the real data. Behaviour depends on how the page is
    opened:
      - Hosted online (e.g. GitHub Pages) at the same URL as the xlsx file:
        on load, it automatically fetches Phuket_Weather.xlsx over HTTP(S)
        (this works fine cross-page on a real web server -- it's only local
        file:// pages that browsers block from reading other local files)
        and a "Refresh" button re-fetches it, so the page always reflects
        whatever the latest scheduled run committed.
      - Opened locally as a plain file (file://): the auto-fetch will fail
        (expected), and it falls back to the manual file-picker / drag-drop
        so you can still point it at a local copy of the xlsx.
    This HTML is identical every time it's generated (no data baked in), so
    generate_html_dashboard() only needs to run once -- see run_cycle().
    """
    config_json = json.dumps(
        {
            "data_sheet_name": DATA_SHEET,
            "xlsx_filename": EXCEL_PATH.name,
            "metrics": [{"code": c, "column": col, "label": label} for c, col, label in METRICS],
            "stations": [
                {"code": c, "label_raw": STATION_LABELS[c], "label_display": STATION_DISPLAY[c]}
                for c in STATION_CODES
            ],
            "rainfall_cap": RAINFALL_CAP_MM,
        }
    )
    html = _HTML_TEMPLATE.replace("__CONFIG_JSON__", config_json)
    html_path.parent.mkdir(parents=True, exist_ok=True)
    html_path.write_text(html, encoding="utf-8")


_HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>Phuket Weather Dashboard</title>
<script src="https://cdn.jsdelivr.net/npm/xlsx@0.18.5/dist/xlsx.full.min.js"></script>
<style>
  :root { --summary-bg: #FFF3E0; --overflow-color: #041C32; }
  body { font-family: Segoe UI, Arial, sans-serif; margin: 24px; background: #fafafa; color: #222; }
  h1 { font-size: 20px; margin-bottom: 4px; }
  .subtitle { color: #666; margin-bottom: 16px; font-size: 13px; }

  .loader { border: 2px dashed #90caf9; border-radius: 8px; padding: 16px 20px; margin-bottom: 18px;
            background: #f4f9ff; display: flex; align-items: center; gap: 14px; flex-wrap: wrap; }
  .loader.dragover { background: #e3f2fd; border-color: #1565C0; }
  .loader button { padding: 8px 14px; font-size: 13px; border-radius: 6px; border: 1px solid #1565C0;
                   background: #1565C0; color: #fff; cursor: pointer; }
  .loader button:hover { background: #0d47a1; }
  #fileStatus { font-size: 13px; color: #555; }
  #fileStatus.error { color: #c62828; font-weight: 600; }
  #fileStatus.ok { color: #2e7d32; }

  .controls { display: flex; gap: 24px; margin-bottom: 16px; align-items: center; }
  .controls label { font-weight: 600; margin-right: 8px; }
  select { padding: 6px 10px; font-size: 14px; border-radius: 6px; border: 1px solid #ccc; }
  table { border-collapse: collapse; font-size: 12px; background: #fff; box-shadow: 0 1px 4px rgba(0,0,0,0.08); }
  th, td { border: 1px solid #ddd; padding: 4px 8px; text-align: center; min-width: 46px; }
  th { background: #f0f0f0; position: sticky; top: 0; }
  td.datecol, th.datecol { position: sticky; left: 0; background: #f0f0f0; font-weight: 600; text-align: left; min-width: 90px; }
  td.summary { background: var(--summary-bg); font-weight: 600; }

  /* Rainfall cells: fixed-height container with a bottom-anchored bar whose
     height represents the value (capped), coloured light -> dark blue. */
  td.rain-cell { position: relative; height: 56px; padding: 2px 4px; vertical-align: bottom; }
  td.rain-cell .bar { position: absolute; left: 4px; right: 4px; bottom: 0; border-radius: 3px 3px 0 0; transition: height 0.15s; }
  td.rain-cell .val { position: relative; z-index: 1; font-weight: 600; text-shadow: 0 0 3px #fff, 0 0 3px #fff; }
  td.rain-cell.overflow .val { color: #fff; text-shadow: none; }

  /* Wind Speed / Land Visibility cells: plain number, no colour or bar. */
  td.plain-cell { height: 56px; vertical-align: middle; }

  .wrap { max-height: 70vh; overflow: auto; border-radius: 6px; }
  .empty { color: #999; }
  .legend { font-size: 12px; color: #555; margin-bottom: 10px; display: flex; align-items: center; gap: 8px; }
  .legend .swatch { display: inline-block; width: 14px; height: 14px; border-radius: 2px; vertical-align: middle; }
  .placeholder { color: #888; font-size: 13px; padding: 20px; }
</style>
</head>
<body>
  <h1>Phuket Weather Dashboard</h1>
  <div class="subtitle">Reads Phuket_Weather.xlsx directly in your browser -- always shows the real, current raw data, never a baked-in snapshot.</div>

  <div id="loader" class="loader">
    <button id="refreshBtn">Refresh from Server</button>
    <button id="pickBtn">Load Local File</button>
    <span id="fileStatus">Loading latest data from the server...</span>
    <input type="file" id="fileInput" accept=".xlsx" style="display:none">
  </div>

  <div class="controls">
    <div><label for="metric">Metric:</label><select id="metric"></select></div>
    <div><label for="station">Station:</label><select id="station"></select></div>
  </div>
  <div id="legend" class="legend"></div>
  <div class="wrap"><table id="grid"></table></div>

<script>
const CONFIG = __CONFIG_JSON__;
const RAIN_CAP = CONFIG.rainfall_cap;

const metricSel = document.getElementById('metric');
const stationSel = document.getElementById('station');
const legendEl = document.getElementById('legend');
const fileInput = document.getElementById('fileInput');
const pickBtn = document.getElementById('pickBtn');
const loaderEl = document.getElementById('loader');
const statusEl = document.getElementById('fileStatus');
const table = document.getElementById('grid');

let DATA = null;

// Dropdowns come from CONFIG (fixed metadata) and are ready immediately,
// before any file is loaded.
CONFIG.metrics.forEach(m => {
  const opt = document.createElement('option');
  opt.value = m.code; opt.textContent = m.label;
  metricSel.appendChild(opt);
});
CONFIG.stations.forEach(s => {
  const opt = document.createElement('option');
  opt.value = s.code; opt.textContent = s.label_display;
  if (s.code === 'COMBINED') opt.selected = true;
  stationSel.appendChild(opt);
});

pickBtn.addEventListener('click', () => fileInput.click());
fileInput.addEventListener('change', e => {
  if (e.target.files && e.target.files[0]) loadFile(e.target.files[0]);
});
['dragover', 'dragenter'].forEach(evt => loaderEl.addEventListener(evt, e => {
  e.preventDefault(); loaderEl.classList.add('dragover');
}));
['dragleave', 'drop'].forEach(evt => loaderEl.addEventListener(evt, e => {
  e.preventDefault(); loaderEl.classList.remove('dragover');
}));
loaderEl.addEventListener('drop', e => {
  if (e.dataTransfer.files && e.dataTransfer.files[0]) loadFile(e.dataTransfer.files[0]);
});

function setStatus(msg, cls) {
  statusEl.textContent = msg;
  statusEl.className = cls || '';
}

function loadFile(file) {
  setStatus('Reading ' + file.name + '...', '');
  const reader = new FileReader();
  reader.onload = evt => {
    try {
      const wb = XLSX.read(evt.target.result, { type: 'array', cellDates: true });
      const sheet = wb.Sheets[CONFIG.data_sheet_name];
      if (!sheet) {
        setStatus(`Couldn't find a "${CONFIG.data_sheet_name}" sheet in ${file.name}.`, 'error');
        return;
      }
      const rows = XLSX.utils.sheet_to_json(sheet, { defval: null });
      DATA = buildDataFromRows(rows);
      const now = new Date();
      setStatus(`Loaded ${file.name} (${rows.length} row(s)) at ${now.toLocaleTimeString()}.`, 'ok');
      render();
    } catch (err) {
      setStatus('Error reading file: ' + err.message, 'error');
    }
  };
  reader.onerror = () => setStatus('Could not read the file.', 'error');
  reader.readAsArrayBuffer(file);
}

function pad(n) { return n.toString().padStart(2, '0'); }

function parseDateTime(v) {
  if (v === null || v === undefined || v === '') return null;
  const d = (v instanceof Date) ? v : new Date(v);
  return isNaN(d.getTime()) ? null : d;
}

// Groups raw rows by Station -> Date -> TimeOfDay -> {column: [values]}, then
// reduces to grid/summary payloads per Metric x Station -- same shape/logic
// as the Excel pivot used to produce, just computed here instead of Python.
// Rows with a blank/unparseable DateTime are skipped (this is the same fix
// as the '<' not supported between float and str bug from the Excel version).
function buildDataFromRows(rows) {
  const buckets = {};   // station -> date -> time -> {col: [values]}
  const dateSet = new Set();
  const timeSet = new Set();
  const columns = CONFIG.metrics.map(m => m.column);

  rows.forEach(row => {
    const dt = parseDateTime(row.DateTime);
    const station = row.Station;
    if (!dt || !station) return;

    const dateStr = `${dt.getFullYear()}-${pad(dt.getMonth() + 1)}-${pad(dt.getDate())}`;
    const timeStr = `${pad(dt.getHours())}:${pad(dt.getMinutes())}`;
    dateSet.add(dateStr); timeSet.add(timeStr);

    buckets[station] = buckets[station] || {};
    buckets[station][dateStr] = buckets[station][dateStr] || {};
    const slot = buckets[station][dateStr][timeStr] =
      buckets[station][dateStr][timeStr] || Object.fromEntries(columns.map(c => [c, []]));

    columns.forEach(col => {
      const v = row[col];
      if (v !== null && v !== undefined && v !== '' && !isNaN(Number(v))) slot[col].push(Number(v));
    });
  });

  const dates = Array.from(dateSet).sort();
  const times = Array.from(timeSet).sort((a, b) => {
    const [ah, am] = a.split(':').map(Number), [bh, bm] = b.split(':').map(Number);
    return (ah * 60 + am) - (bh * 60 + bm);
  });

  const round2 = x => Math.round(x * 100) / 100;
  const mean = arr => arr.length ? round2(arr.reduce((a, b) => a + b, 0) / arr.length) : null;

  const payload = {};
  CONFIG.metrics.forEach(m => {
    payload[m.code] = {};
    CONFIG.stations.forEach(s => {
      const stationLabel = s.label_raw;
      const grid = dates.map(d => times.map(t => {
        const slot = buckets[stationLabel] && buckets[stationLabel][d] && buckets[stationLabel][d][t];
        return slot ? mean(slot[m.column]) : null;
      }));
      const summary = grid.map(rowVals => {
        const present = rowVals.filter(v => v !== null);
        if (!present.length) return null;
        return m.code === 'Rain' ? round2(present.reduce((a, b) => a + b, 0)) : mean(present);
      });
      payload[m.code][s.code] = { grid, summary };
    });
  });

  return { dates, times, payload };
}

function rainColor(ratio) {
  // light blue -> dark blue, ratio in [0,1]
  const c1 = [227, 242, 253], c2 = [21, 101, 192];
  const lerp = (a, b) => Math.round(a + (b - a) * ratio);
  return `rgb(${lerp(c1[0], c2[0])}, ${lerp(c1[1], c2[1])}, ${lerp(c1[2], c2[2])})`;
}

function renderLegend(metric) {
  legendEl.innerHTML = '';
  if (metric !== 'Rain') return;
  const items = [
    [0, 'Light blue (0mm)'],
    [0.5, 'Mid'],
    [1, `Dark blue (${RAIN_CAP}mm)`],
  ];
  items.forEach(([ratio, label]) => {
    const sw = document.createElement('span');
    sw.className = 'swatch';
    sw.style.background = rainColor(ratio);
    legendEl.appendChild(sw);
    legendEl.appendChild(document.createTextNode(label + '  '));
  });
  const overflowSw = document.createElement('span');
  overflowSw.className = 'swatch';
  overflowSw.style.background = 'var(--overflow-color)';
  legendEl.appendChild(overflowSw);
  legendEl.appendChild(document.createTextNode(`Above ${RAIN_CAP}mm (very dark)`));
}

function render() {
  if (!DATA) {
    table.innerHTML = '<tr><td class="placeholder">Load Phuket_Weather.xlsx above to see the dashboard.</td></tr>';
    return;
  }

  const metric = metricSel.value;
  const station = stationSel.value;
  const combo = DATA.payload[metric][station];
  const grid = combo.grid, summary = combo.summary;
  const isRain = metric === 'Rain';

  renderLegend(metric);

  table.innerHTML = '';

  const thead = document.createElement('tr');
  thead.appendChild(Object.assign(document.createElement('th'), {textContent: 'Date', className: 'datecol'}));
  DATA.times.forEach(t => thead.appendChild(Object.assign(document.createElement('th'), {textContent: t})));
  const summaryTh = document.createElement('th');
  summaryTh.textContent = 'Daily Summary';
  thead.appendChild(summaryTh);
  table.appendChild(thead);

  DATA.dates.forEach((d, i) => {
    const tr = document.createElement('tr');
    const dateTd = document.createElement('td');
    dateTd.textContent = d; dateTd.className = 'datecol';
    tr.appendChild(dateTd);

    grid[i].forEach(v => {
      const td = document.createElement('td');

      if (isRain) {
        td.className = 'rain-cell';
        if (v !== null) {
          const overflow = v > RAIN_CAP;
          const ratio = Math.min(v, RAIN_CAP) / RAIN_CAP;
          const bar = document.createElement('div');
          bar.className = 'bar';
          bar.style.height = (ratio * 100) + '%';
          bar.style.background = overflow ? 'var(--overflow-color)' : rainColor(ratio);
          if (overflow) td.classList.add('overflow');
          td.appendChild(bar);
        } else {
          td.classList.add('empty');
        }
        const span = document.createElement('span');
        span.className = 'val';
        span.textContent = v === null ? '' : v;
        td.appendChild(span);
      } else {
        // Wind Speed / Land Visibility: plain number only, no colour/bar.
        td.className = 'plain-cell';
        td.textContent = v === null ? '' : v;
        if (v === null) td.classList.add('empty');
      }

      tr.appendChild(td);
    });

    const sumTd = document.createElement('td');
    sumTd.className = 'summary';
    sumTd.textContent = summary[i] === null ? '' : summary[i];
    tr.appendChild(sumTd);

    table.appendChild(tr);
  });
}

metricSel.addEventListener('change', render);
stationSel.addEventListener('change', render);
render();
</script>
</body>
</html>
"""


# ----------------------------------------------------------------------------
# 4. Orchestration
# ----------------------------------------------------------------------------

def run_cycle():
    print(f"[{datetime.now().isoformat()}] Fetching TMD feed...")
    records = fetch_records()  # AIRPORT + PHUKET raw rows, plus a COMBINED row, per timestamp
    print(f"Built {len(records)} row(s) (raw + combined) this cycle.")

    added, total = update_excel(records)
    print(f"Excel Data sheet updated (raw data only): +{added} new row(s), {total} total.")

    # The HTML dashboard reads Phuket_Weather.xlsx LIVE in the browser (see
    # generate_html_dashboard's docstring), so it never needs regenerating
    # with new data -- this call is only here to (re)create the file if it's
    # ever missing. Safe/cheap to leave in every cycle.
    if not HTML_PATH.exists():
        generate_html_dashboard(HTML_PATH)
        print(f"HTML dashboard created. Saved to {HTML_PATH}")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--once", action="store_true", help="Run a single cycle and exit")
    args = parser.parse_args()

    if args.once:
        run_cycle()
        return

    while True:
        try:
            run_cycle()
        except Exception as e:
            print(f"Error this cycle: {e}")
        print(f"Sleeping {INTERVAL_HOURS} hour(s)...")
        time.sleep(INTERVAL_HOURS * 3600)


if __name__ == "__main__":
    main()